from flask import Blueprint, jsonify, request, current_app, send_file
from flask_jwt_extended import jwt_required
from sqlalchemy import text
from datetime import datetime
from dateutil.relativedelta import relativedelta
from ..extensions import db
from .dashboard_common import get_date_filter
from decimal import Decimal
import io
import csv


# --------------------------------------------------------
#   BLUEPRINT — ALL ENDPOINTS IN THIS FILE
# --------------------------------------------------------
bp = Blueprint("taxpayer_report_risk_profiling", __name__, url_prefix="/api/taxpayer_report_risk_profiling")


# --------------------------------------------------------
#   HELPER FUNCTION
# --------------------------------------------------------
def get_range_dates():
    """
    Converts range_type (1m,3m,6m,1y,custom) into numeric (start_year, start_month, end_year, end_month)
    """
    range_type = request.args.get("range_type") or "custom"
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    if start_date and end_date:
        try:
            sd = datetime.strptime(start_date, "%Y-%m-%d")
            ed = datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            try:
                sd = datetime.strptime(start_date, "%d-%m-%Y")
                ed = datetime.strptime(end_date, "%d-%m-%Y")
            except ValueError:
                sd = None
                ed = None
        if sd and ed:
            return sd.year, sd.month, ed.year, ed.month

    today = datetime.today()

    if range_type == "1m":
        sd = today.replace(day=1)

    elif range_type == "3m":
        sd = (today - relativedelta(months=2)).replace(day=1)

    elif range_type == "6m":
        sd = (today - relativedelta(months=5)).replace(day=1)

    elif range_type == "1y":
        sd = (today - relativedelta(years=1)).replace(day=1)

    else:
        sd = today.replace(day=1)

    return sd.year, sd.month, today.year, today.month

def _safe_str(value):
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        return v if v else None
    return value

def _to_display(value):
    """
    Display formatting rules:
    - None/empty -> "NA"
    - 0 -> "NA" (display only)
    """
    if value is None:
        return "NA"
    if isinstance(value, str):
        v = value.strip()
        return v if v else "NA"
    if isinstance(value, (int, float, Decimal)):
        try:
            if float(value) == 0:
                return "NA"
        except (TypeError, ValueError):
            pass
        return float(value) if isinstance(value, Decimal) else value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value

def _join_non_empty(parts, sep=", "):
    items = []
    for p in parts:
        v = _safe_str(p)
        if v:
            items.append(v)
    return sep.join(items) if items else None

def _is_no_data(structured_report):
    if not structured_report:
        return True
    return all((r.get("value") in ("NA", "", None)) for r in structured_report)

def _get_structured_report_from_summary():
    result = get_summary()
    if isinstance(result, tuple):
        resp = result[0]
        status = result[1]
        data = resp.get_json() if hasattr(resp, "get_json") else None
        if status != 200:
            return {"_error": True, "status_code": status, "payload": data or {"error": "Unknown error"}}
    else:
        resp = result
        data = resp.get_json() if hasattr(resp, "get_json") else None
    return data or {}

def _risk_fill(value):
    if not value:
        return "NA"
    v = str(value).strip().lower()
    if v == "high":
        return "HIGH"
    if v == "medium":
        return "MEDIUM"
    if v == "low":
        return "LOW"
    return "NA"


def _pdf_escape(value):
    return str(value).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _build_simple_pdf_bytes(lines):
    page_width = 595
    page_height = 842
    left = 40
    top = 800
    line_height = 14
    lines_per_page = 50

    chunks = [lines[i:i + lines_per_page] for i in range(0, len(lines), lines_per_page)] or [["No data"]]

    objects = []

    def add_object(content: bytes) -> int:
        objects.append(content)
        return len(objects)

    font_obj = add_object(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")

    page_obj_ids = []
    content_obj_ids = []

    for chunk in chunks:
        text_lines = [
            "BT",
            "/F1 10 Tf",
            f"{left} {top} Td",
            f"{line_height} TL",
        ]
        for line in chunk:
            text_lines.append(f"({_pdf_escape(line)}) Tj")
            text_lines.append("T*")
        text_lines.append("ET")
        stream = "\n".join(text_lines).encode("latin-1", errors="replace")
        content_obj = add_object(
            f"<< /Length {len(stream)} >>\nstream\n".encode("latin-1") +
            stream +
            b"\nendstream"
        )
        content_obj_ids.append(content_obj)
        page_obj_ids.append(add_object(b""))

    kids = " ".join(f"{obj_id} 0 R" for obj_id in page_obj_ids)
    pages_obj = add_object(
        f"<< /Type /Pages /Count {len(page_obj_ids)} /Kids [{kids}] >>".encode("latin-1")
    )

    for index, page_obj_id in enumerate(page_obj_ids):
        page_content = (
            f"<< /Type /Page /Parent {pages_obj} 0 R "
            f"/MediaBox [0 0 {page_width} {page_height}] "
            f"/Resources << /Font << /F1 {font_obj} 0 R >> >> "
            f"/Contents {content_obj_ids[index]} 0 R >>"
        ).encode("latin-1")
        objects[page_obj_id - 1] = page_content

    catalog_obj = add_object(f"<< /Type /Catalog /Pages {pages_obj} 0 R >>".encode("latin-1"))

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{idx} 0 obj\n".encode("latin-1"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")

    xref_offset = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("latin-1"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("latin-1"))
    pdf.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root {catalog_obj} 0 R >>\nstartxref\n{xref_offset}\n%%EOF".encode("latin-1")
    )
    return bytes(pdf)

def _table_exists(table_name: str) -> bool:
    try:
        row = db.session.execute(
            text(
                """
                SELECT COUNT(*) AS c
                FROM information_schema.tables
                WHERE table_schema = DATABASE()
                  AND table_name = :t
                """
            ),
            {"t": table_name},
        ).fetchone()
        return bool(row and int(row[0] or 0) > 0)
    except Exception:
        return False

def _table_has_column(table_name: str, column_name: str) -> bool:
    try:
        row = db.session.execute(
            text(
                """
                SELECT COUNT(*) AS c
                FROM information_schema.columns
                WHERE table_schema = DATABASE()
                  AND table_name = :t
                  AND column_name = :c
                """
            ),
            {"t": table_name, "c": column_name},
        ).fetchone()
        return bool(row and int(row[0] or 0) > 0)
    except Exception:
        return False


# --------------------------------------------------------
#   API 1: GET DISTINCT TIN
# --------------------------------------------------------
@bp.route("/dropdown", methods=["OPTIONS"])
def taxpayer_dropdown_options():
    return ("", 200)


@bp.get("/dropdown")
@jwt_required()
def taxpayer_dropdown():
    try:
        search = (request.args.get("q") or "").strip()

        params = {}
        where = "WHERE trm.tin IS NOT NULL"
        if search:
            params["s"] = f"%{search}%"
            where += """
                AND (
                    CAST(trm.tin AS CHAR(30)) LIKE :s
                    OR trm.taxpayername LIKE :s
                    OR trm.maintradename LIKE :s
                )
            """

        query = f"""
            SELECT
                trm.tin AS tin,
                COALESCE(
                    NULLIF(TRIM(trm.taxpayername), ''),
                    NULLIF(TRIM(trm.maintradename), ''),
                    'Unknown'
                ) AS taxpayer_name
            FROM tin_registration_mst trm
            {where}
            ORDER BY taxpayer_name ASC
            LIMIT 100
        """

        rows = db.session.execute(text(query), params).fetchall()
        return jsonify(
            [
                {
                    "value": r._mapping["tin"],
                    "label": f'{r._mapping["tin"]} - {r._mapping["taxpayer_name"]}',
                    "name": r._mapping["taxpayer_name"],
                    "tin": r._mapping["tin"],
                    "taxpayer_name": r._mapping["taxpayer_name"],
                }
                for r in rows
                if r._mapping.get("tin") and r._mapping.get("taxpayer_name")
            ]
        ), 200
    except Exception as e:
        current_app.logger.exception(e)
        return jsonify([]), 200

# --------------------------------------------------------
#   API 2: GET TAX COMPLIANCE SUMMARY TABLE + NEW FIELDS
# --------------------------------------------------------
@bp.route("/taxpayer-summary", methods=["OPTIONS"])
def get_summary_options():
    return ("", 200)


@bp.get("/taxpayer-summary")
@jwt_required()
def get_summary():
    try:
        tin = request.args.get("tin")

        if not tin:
            return jsonify({"error": "TIN is required"}), 400

        start_y, start_m, end_y, end_m = get_range_dates()
        if None in (start_y, start_m, end_y, end_m):
            today = datetime.today()
            start_y, start_m, end_y, end_m = today.year, today.month, today.year, today.month
        tin_norm = tin.replace(" ", "").strip().upper()

        # ==========================
        #  GST SECTION
        # ==========================

        gst_date_filter, gst_date_params = get_date_filter(
            column_year="pr.tax_period_year",

        )
        where_period = f"""
            TRIM(CAST(pr.tin AS CHAR(20))) = TRIM(:tin)
            AND ({gst_date_filter})
        """

        params = {"tin": tin_norm, **gst_date_params}

        # ---------------- GST Overview ----------------
        overview = dict(db.session.execute(text(f"""
            SELECT
              COUNT(*) AS total_records,
              COUNT(DISTINCT CONCAT(pr.tax_period_year,'-',pr.tax_period_month)) AS total_months_filed,
              COALESCE(SUM(COALESCE(pr.input_credits, 0)), 0) AS total_input_credit,
              COALESCE(SUM(COALESCE(pr.output_debits, 0)), 0) AS total_output_debit,
              COALESCE(SUM(COALESCE(pr.gst_payable, 0) - COALESCE(pr.gst_refundable, 0)), 0) AS total_net_tax,
              COALESCE(AVG(COALESCE(pr.gst_payable, 0) - COALESCE(pr.gst_refundable, 0)), 0) AS average_monthly_tax
            FROM gst_fraud_justification pr
            WHERE {where_period}
        """), params).fetchone()._mapping)

        # ---------------- GST Payable / Refundable ----------------
        payable = dict(db.session.execute(text(f"""
            SELECT 
              COALESCE(SUM(COALESCE(pr.gst_payable, 0)), 0) AS total_payable,
              COALESCE(SUM(COALESCE(pr.gst_refundable, 0)), 0) AS total_refundable
            FROM gst_fraud_justification pr
            WHERE {where_period}
        """), params).fetchone()._mapping)

        # ---------------- GST Input / Output ----------------
        input_output = dict(db.session.execute(text(f"""
            SELECT 
              COALESCE(SUM(COALESCE(pr.input_credits, 0)), 0) AS total_input_credit,
              COALESCE(SUM(COALESCE(pr.output_debits, 0)), 0) AS total_output_debit
            FROM gst_fraud_justification pr
            WHERE {where_period}
        """), params).fetchone()._mapping)

        # ----------------  GST Compliance Metrics (FIXED) ----------------
        gst_compliance = dict(db.session.execute(text(f"""
            SELECT
                SUM(
                    CASE
                        WHEN received_date IS NOT NULL AND due_date IS NOT NULL
                        AND STR_TO_DATE(received_date, '%Y-%m-%d') >
                            STR_TO_DATE(due_date, '%Y-%m-%d')
                        THEN 1 ELSE 0
                    END
                ) AS payment_delay_count,

                AVG(
                    CASE
                        WHEN received_date IS NOT NULL AND due_date IS NOT NULL
                        THEN DATEDIFF(
                            STR_TO_DATE(received_date, '%Y-%m-%d'),
                            STR_TO_DATE(due_date, '%Y-%m-%d')
                        )
                    END
                ) AS average_delay_days

            FROM gst_fraud_justification pr
            WHERE {where_period}
        """), params).fetchone()._mapping)

        # ---------------- GST Fraud Summary ----------------
        fraud = dict(db.session.execute(text(f"""
            SELECT
                SUM(CASE WHEN COALESCE(pr.is_fraud, 0) = 1 THEN 1 ELSE 0 END) AS total_fraud_cases,
                ROUND(
                    (
                        SUM(CASE WHEN COALESCE(pr.is_fraud, 0) = 1 THEN 1 ELSE 0 END) * 100.0
                    ) / NULLIF(COUNT(*), 0),
                2) AS fraud_percentage,
                '' AS fraud_reasons
            FROM gst_fraud_justification pr
            WHERE {where_period}
        """), params).fetchone()._mapping)

        ratio = round(
            (input_output.get("total_input_credit", 0) or 0) /
            (input_output.get("total_output_debit", 0) or 1), 2
        )

        gst_response = {
            "overview": overview,
            "payable_vs_refundable": payable,
            "input_vs_output": {
                **input_output,
                "ratio_input_output": ratio
            },
            "compliance_metrics": {
                "payment_delay_count": gst_compliance.get("payment_delay_count", 0),
                "average_delay_days": gst_compliance.get("average_delay_days", 0)
            },
            "fraud_summary": fraud
        }

        # ==========================
        # SWT SECTION
        # ==========================

        swt_date_filter, swt_date_params = get_date_filter(
            column_year="psr.tax_period_year",

        )
        swt_where = f"""
            TRIM(CAST(psr.tin AS CHAR(20))) = TRIM(:tin)
            AND ({swt_date_filter})
        """
        swt_params = {"tin": tin_norm, **swt_date_params}

        swt_overview = db.session.execute(text(f"""
            SELECT
                COUNT(*) AS total_transactions,
                0 AS risk_score
            FROM swt_fraud_justification psr
            WHERE {swt_where}
        """), swt_params).fetchone()

        gst_fraud = fraud.get("total_fraud_cases", 0) or 0
        swt_fraud = db.session.execute(text(f"""
            SELECT
                SUM(CASE WHEN LOWER(COALESCE(psr.predicted_fraud,'')) = 'fraud' THEN 1 ELSE 0 END) AS c
            FROM swt_fraud_justification psr
            WHERE {swt_where}
        """), swt_params).scalar()
        swt_fraud = swt_fraud or 0

        gst_reason_rows = db.session.execute(text(f"""
            SELECT explanation
            FROM gst_fraud_justification pr
            WHERE {where_period}
        """), params).fetchall()

        swt_reason_rows = db.session.execute(text(f"""
            SELECT explanation
            FROM swt_fraud_justification psr
            WHERE {swt_where}
        """), swt_params).fetchall()

        def extract_reasons(rows):
            import re
            from collections import Counter
            regex = re.compile(r"^\s*(\d+)\.\s*([^\n:]+)", re.MULTILINE)
            bag = Counter()
            for (logic,) in rows:
                for m in regex.finditer(logic or ""):
                    bag[m.group(2).strip()] += 1
            return [k for k, v in bag.most_common(5)]

        swt_risk_level = "Unknown"
        try:
            total_tx = int(getattr(swt_overview, "total_transactions", 0) or 0)
            if total_tx > 0:
                swt_risk_level = "High" if int(swt_fraud or 0) > 0 else "Low"
        except Exception:
            pass

        swt_response = {
            "overview": {
                "risk_level": swt_risk_level,
                "risk_score": getattr(swt_overview, "risk_score", 0),
                "risk_factors": swt_fraud,
                "total_transactions": getattr(swt_overview, "total_transactions", 0)
            },
            "fraud_metrics": {
                "total_fraud_cases": gst_fraud + swt_fraud,
                "fraud_percentage": round((swt_fraud / (getattr(swt_overview, "total_transactions", 0) or 1)) * 100, 2),
                "breakdown": {
                    "gst_fraud_cases": gst_fraud,
                    "swt_fraud_cases": swt_fraud
                },
                "fraud_patterns": {
                    "gst_reasons": extract_reasons(gst_reason_rows),
                    "swt_reasons": extract_reasons(swt_reason_rows)
                }
            },
            "compliance_metrics": {
                "total_delayed_filings": 0,
                "delay_percentage": 0,
                "breakdown": {"gst_delays": 0, "swt_delays": 0}
            }
        }

        # ==========================
        # FINAL RESPONSE
        # ==========================

        # ==========================
        # STRUCTURED REPORT (NEW)
        # ==========================
        # --- Taxpayer Registration Details ---
        trm = db.session.execute(text("""
            SELECT
                taxcentre, tin, taxpayertype, taxpayername, maintradename,
                enterprisetype, entstartdate,
                individualsituation,
                pitaccountno, citaccountno, swtaccountno, gstaccountno,
                iwtaccountno, mfwtaccountno, fcwtaccountno,
                physicaladdressprovince, mailingaddressprovince,
                address1, address2, address3, address4,
                physicaladdress5, physicaladdress6, physicaladdress7, physicaladdress8,
                city, province,
                sectoractivity, enterpriseactivity, entactivitycode,
                entcontactname, title, phone1, phone2, entcontemail,
                contactname, contactphone, contactmobile, contactemail,
                repcontactname, reptype, repphone1, repphone2, repcontemail
            FROM tin_registration_mst
            WHERE TRIM(tin) = TRIM(:tin)
               OR normalized_tin = :tin_norm
            LIMIT 1
        """), {"tin": tin, "tin_norm": tin_norm}).fetchone()

        trm_map = trm._mapping if trm else {}

        # Taxpayer segmentation (latest GST record)
        gst_segment_row = db.session.execute(text("""
            SELECT pr.taxpayer_type
            FROM gst_fraud_justification pr
            WHERE TRIM(CAST(pr.tin AS CHAR(20))) = TRIM(:tin)
            ORDER BY pr.tax_period_year DESC, pr.tax_period_month DESC, pr.id DESC
            LIMIT 1
        """), {"tin": tin_norm}).fetchone()
        gst_segmentation = gst_segment_row[0] if gst_segment_row else None

        # --- GST/SWT/CIT Compliance Metrics ---
        expected_months = ((end_y - start_y) * 12) + (end_m - start_m) + 1

        def build_expected_months(sy, sm, ey, em):
            expected = set()
            for y in range(sy, ey + 1):
                m_start = sm if y == sy else 1
                m_end = em if y == ey else 12
                for m in range(m_start, m_end + 1):
                    expected.add((y, m))
            return expected

        gst_months_row = db.session.execute(text(f"""
            SELECT
                COUNT(*) AS total_records,
                COUNT(DISTINCT CONCAT(pr.tax_period_year,'-',LPAD(pr.tax_period_month,2,'0'))) AS months_reported,
                COALESCE(SUM(pr.gst_payable), 0) AS sum_gst_payable,
                COALESCE(SUM(pr.input_credits), 0) AS sum_input_credits,
                COALESCE(SUM(pr.output_debits), 0) AS sum_output_debits,
                COALESCE(SUM(pr.total_sales_income), 0) AS sum_sales
            FROM gst_fraud_justification pr
            WHERE {where_period}
        """), params).fetchone()._mapping

        swt_months_row = db.session.execute(text("""
            SELECT
                COUNT(*) AS total_records,
                COUNT(DISTINCT CONCAT(pr.tax_period_year,'-',LPAD(pr.tax_period_month,2,'0'))) AS months_reported,
                COALESCE(SUM(pr.total_swt_tax_deducted), 0) AS sum_swt_deducted,
                COALESCE(MAX(pr.tax_period_year*100 + pr.tax_period_month), 0) AS max_period
            FROM swt_fraud_justification pr
            WHERE TRIM(CAST(pr.tin AS CHAR(20))) = TRIM(:tin_norm)
              AND (pr.tax_period_year > :sy OR (pr.tax_period_year = :sy AND pr.tax_period_month >= :sm))
              AND (pr.tax_period_year < :ey OR (pr.tax_period_year = :ey AND pr.tax_period_month <= :em))
        """), {"tin_norm": tin_norm, "sy": start_y, "sm": start_m, "ey": end_y, "em": end_m}).fetchone()._mapping

        cit_years_row = db.session.execute(text("""
            SELECT
                COUNT(*) AS total_records,
                COUNT(DISTINCT c.tax_period_year) AS years_reported,
                COALESCE(SUM(c.total_tax_payable), 0) AS sum_tax_payable,
                MAX(c.tax_period_year) AS max_year
            FROM cit_fraud_justification c
            WHERE TRIM(c.tin) = TRIM(:tin)
              AND c.tax_period_year BETWEEN :sy AND :ey
        """), {"tin": tin, "sy": start_y, "ey": end_y}).fetchone()._mapping

        # Latest SWT employees (fallback to previous valid)
        swt_latest = db.session.execute(text("""
            SELECT employees_on_payroll, employees_paid_swt
            FROM swt_fraud_justification
            WHERE TRIM(CAST(tin AS CHAR(20))) = TRIM(:tin_norm)
              AND (employees_on_payroll IS NOT NULL OR employees_paid_swt IS NOT NULL)
            ORDER BY tax_period_year DESC, tax_period_month DESC
            LIMIT 1
        """), {"tin_norm": tin_norm}).fetchone()

        # Latest CIT assets/liabilities (max year with data)
        cit_max_year_row = db.session.execute(text("""
            SELECT MAX(tax_period_year) AS max_year
            FROM cit_fraud_justification
            WHERE TRIM(tin) = TRIM(:tin)
              AND tax_period_year BETWEEN :sy AND :ey
        """), {"tin": tin, "sy": start_y, "ey": end_y}).fetchone()
        cit_max_year = cit_max_year_row[0] if cit_max_year_row else None
        cit_latest = None
        if cit_max_year:
            if _table_has_column("cit_fraud_justification", "total_assets") and _table_has_column("cit_fraud_justification", "total_liabilities"):
                cit_latest = db.session.execute(text("""
                    SELECT total_assets, total_liabilities
                    FROM cit_fraud_justification
                    WHERE TRIM(tin) = TRIM(:tin)
                      AND tax_period_year = :yy
                    ORDER BY tax_period_year DESC, id DESC
                    LIMIT 1
                """), {"tin": tin, "yy": cit_max_year}).fetchone()
            cit_year_tax_row = db.session.execute(text("""
                SELECT COALESCE(SUM(total_tax_payable), 0) AS sum_tax_payable
                FROM cit_fraud_justification
                WHERE TRIM(tin) = TRIM(:tin)
                  AND tax_period_year = :yy
            """), {"tin": tin, "yy": cit_max_year}).fetchone()
        else:
            cit_year_tax_row = None

        # Risk analysis inputs
        cit_risk_row = db.session.execute(text("""
            SELECT
                COUNT(*) AS total_records,
                SUM(CASE WHEN LOWER(predicted_fraud) = 'fraud' THEN 1 ELSE 0 END) AS fraud_count
            FROM cit_fraud_justification
            WHERE TRIM(tin) = TRIM(:tin)
              AND tax_period_year BETWEEN :sy AND :ey
        """), {"tin": tin, "sy": start_y, "ey": end_y}).fetchone()._mapping

        gst_risk_row = db.session.execute(text(f"""
            SELECT
                COUNT(*) AS total_records,
                SUM(CASE WHEN COALESCE(pr.is_fraud, 0) = 1 THEN 1 ELSE 0 END) AS fraud_count
            FROM gst_fraud_justification pr
            WHERE {where_period}
        """), params).fetchone()._mapping

        swt_risk_row = db.session.execute(text("""
            SELECT
                COUNT(*) AS total_records,
                SUM(CASE WHEN LOWER(COALESCE(predicted_fraud,'')) = 'fraud' THEN 1 ELSE 0 END) AS flag_count
            FROM swt_fraud_justification
            WHERE TRIM(CAST(tin AS CHAR(20))) = TRIM(:tin_norm)
              AND (tax_period_year > :sy OR (tax_period_year = :sy AND tax_period_month >= :sm))
              AND (tax_period_year < :ey OR (tax_period_year = :ey AND tax_period_month <= :em))
        """), {"tin_norm": tin_norm, "sy": start_y, "sm": start_m, "ey": end_y, "em": end_m}).fetchone()._mapping

        # Derived calculations
        gst_total_records = int(gst_months_row.get("total_records") or 0)
        swt_total_records = int(swt_months_row.get("total_records") or 0)
        cit_total_records = int(cit_years_row.get("total_records") or 0)

        gst_months_reported = int(gst_months_row.get("months_reported") or 0)
        swt_months_reported = int(swt_months_row.get("months_reported") or 0)
        cit_years_reported = int(cit_years_row.get("years_reported") or 0)
        expected_years = (end_y - start_y) + 1

        gst_outstanding = None
        swt_outstanding = None
        cit_outstanding = None

        expected_set = build_expected_months(start_y, start_m, end_y, end_m)
        gst_actual_rows = []
        swt_actual_rows = []
        cit_year_rows = []

        if gst_total_records:
            gst_actual_rows = db.session.execute(text(f"""
                SELECT DISTINCT pr.tax_period_year AS y, pr.tax_period_month AS m
                FROM gst_fraud_justification pr
                WHERE {where_period}
            """), params).fetchall()
            gst_actual = {(int(r.y), int(r.m)) for r in gst_actual_rows if r.y is not None and r.m is not None}
            gst_outstanding = len(expected_set - gst_actual)

        if swt_total_records:
            swt_actual_rows = db.session.execute(text("""
                SELECT DISTINCT pr.tax_period_year AS y, pr.tax_period_month AS m
                FROM swt_fraud_justification pr
                WHERE TRIM(CAST(pr.tin AS CHAR(20))) = TRIM(:tin_norm)
                  AND (pr.tax_period_year > :sy OR (pr.tax_period_year = :sy AND pr.tax_period_month >= :sm))
                  AND (pr.tax_period_year < :ey OR (pr.tax_period_year = :ey AND pr.tax_period_month <= :em))
            """), {"tin_norm": tin_norm, "sy": start_y, "sm": start_m, "ey": end_y, "em": end_m}).fetchall()
            swt_actual = {(int(r.y), int(r.m)) for r in swt_actual_rows if r.y is not None and r.m is not None}
            swt_outstanding = len(expected_set - swt_actual)

        if cit_total_records:
            cit_year_rows = db.session.execute(text("""
                SELECT DISTINCT c.tax_period_year AS y
                FROM cit_fraud_justification c
                WHERE TRIM(c.tin) = TRIM(:tin)
                  AND c.tax_period_year BETWEEN :sy AND :ey
            """), {"tin": tin, "sy": start_y, "ey": end_y}).fetchall()
            cit_years = {int(r.y) for r in cit_year_rows if r.y is not None}
            expected_years_set = set(range(start_y, end_y + 1))
            cit_outstanding = len(expected_years_set - cit_years)

        sum_output_debits = gst_months_row.get("sum_output_debits")
        sum_input_credits = gst_months_row.get("sum_input_credits")
        sum_gst_payable = gst_months_row.get("sum_gst_payable")
        if sum_output_debits is None or (sum_output_debits == 0 and (sum_gst_payable or 0) != 0):
            gst_account_balance = (sum_gst_payable or 0) - (sum_input_credits or 0)
        else:
            gst_account_balance = (sum_output_debits or 0) - (sum_input_credits or 0)

        # Internal GST consistency check (do not expose)
        primary_balance = (sum_output_debits or 0) - (sum_input_credits or 0)
        fallback_balance = (sum_gst_payable or 0) - (sum_input_credits or 0)
        denom = abs(fallback_balance) if abs(fallback_balance) > 0 else 1
        gst_mismatch_pct = (abs(primary_balance - fallback_balance) / denom) * 100
        if gst_mismatch_pct > 5:
            try:
                current_app.logger.warning(
                    "GST balance mismatch >5%% for tin=%s (primary=%s, fallback=%s, pct=%.2f)",
                    tin, primary_balance, fallback_balance, gst_mismatch_pct
                )
            except Exception:
                pass
        swt_account_balance = swt_months_row.get("sum_swt_deducted")
        cit_account_balance = cit_year_tax_row[0] if cit_year_tax_row else None

        # Employees (latest SWT)
        employees_val = None
        if swt_latest:
            employees_val = swt_latest[0] if swt_latest[0] is not None else swt_latest[1]

        # Annual Turnover (GST sum)
        annual_turnover = gst_months_row.get("sum_sales")

        # Assets & Liabilities (CIT latest)
        total_assets = cit_latest[0] if cit_latest else None
        total_liabilities = cit_latest[1] if cit_latest else None

        # Risks
        cit_risk = None
        if cit_total_records == 0:
            cit_risk = None
        else:
            if (cit_risk_row.get("total_records") or 0) == 0:
                cit_risk = None
            else:
                cit_risk = "High" if (cit_risk_row.get("fraud_count") or 0) > 0 else "Low"

        gst_risk = None
        if (gst_risk_row.get("total_records") or 0) == 0:
            gst_risk = None
        else:
            gst_risk = "High" if (gst_risk_row.get("fraud_count") or 0) > 0 else "Low"

        swt_risk = None
        if (swt_risk_row.get("total_records") or 0) == 0:
            swt_risk = None
        else:
            swt_risk = "High" if (swt_risk_row.get("flag_count") or 0) > 0 else "Low"

        # Overall risk priority: High > Medium > Low > NA
        overall_risk = None
        risk_values = [cit_risk, gst_risk, swt_risk]
        if any(r == "High" for r in risk_values):
            overall_risk = "High"
        elif any(r == "Medium" for r in risk_values):
            overall_risk = "Medium"
        elif any(r == "Low" for r in risk_values):
            overall_risk = "Low"

        # Recommendations
        cit_reco = "Audit Required" if cit_risk == "High" else ("No Action" if cit_risk == "Low" else None)
        gst_reco = "Investigate" if gst_risk == "High" else ("Normal" if gst_risk == "Low" else None)
        swt_reco = "Verify Payroll" if swt_risk == "High" else ("Normal" if swt_risk == "Low" else None)
        other_reco = None

        other_accounts = _join_non_empty([
            f"IWT: {trm_map.get('iwtaccountno')}" if _safe_str(trm_map.get("iwtaccountno")) else None,
            f"MFWT: {trm_map.get('mfwtaccountno')}" if _safe_str(trm_map.get("mfwtaccountno")) else None,
            f"FCWT: {trm_map.get('fcwtaccountno')}" if _safe_str(trm_map.get("fcwtaccountno")) else None,
        ])

        individual_situation = _safe_str(trm_map.get("individualsituation"))
        sbt_a = "Yes" if individual_situation and "SBT-A" in individual_situation.upper() else None
        sbt_q = "Yes" if individual_situation and "SBT-Q" in individual_situation.upper() else None

        residential_address = _join_non_empty([
            trm_map.get("physicaladdress5"),
            trm_map.get("physicaladdress6"),
            trm_map.get("physicaladdress7"),
            trm_map.get("physicaladdress8"),
            trm_map.get("city"),
            trm_map.get("province"),
        ])

        mailing_address = _join_non_empty([
            trm_map.get("address1"),
            trm_map.get("address2"),
            trm_map.get("address3"),
            trm_map.get("address4"),
            trm_map.get("mailingaddressprovince"),
        ])

        phone_number = _safe_str(trm_map.get("phone1")) or _safe_str(trm_map.get("phone2")) or _safe_str(trm_map.get("contactphone")) or _safe_str(trm_map.get("contactmobile"))
        rep_phone = _safe_str(trm_map.get("repphone1")) or _safe_str(trm_map.get("repphone2"))

        structured_report = [
            {"label": "TAXPAYER DETAILS", "value": ""},
            {"label": "TIN", "value": _to_display(trm_map.get("tin") or tin)},
            {"label": "TaxpayerName", "value": _to_display(trm_map.get("taxpayername"))},
            {"label": "MainTradeName", "value": _to_display(trm_map.get("maintradename"))},
            {"label": "Taxpayer Segmentation", "value": _to_display(gst_segmentation)},
            {"label": "Tax Centre", "value": _to_display(trm_map.get("taxcentre"))},
            {"label": "TaxpayerType", "value": _to_display(trm_map.get("taxpayertype"))},
            {"label": "EnterpriseType", "value": _to_display(trm_map.get("enterprisetype"))},
            {"label": "Start Date", "value": _to_display(trm_map.get("entstartdate"))},

            {"label": "TAX ACCOUNT DETAILS", "value": ""},
            {"label": "PITTaxAccount", "value": _to_display(trm_map.get("pitaccountno"))},
            {"label": "CITTaxAccount", "value": _to_display(trm_map.get("citaccountno"))},
            {"label": "SWTTaxAccount", "value": _to_display(trm_map.get("swtaccountno"))},
            {"label": "GSTTaxAccount", "value": _to_display(trm_map.get("gstaccountno"))},
            {"label": "SBT-ATaxAccount", "value": _to_display(sbt_a)},
            {"label": "SBT-QTaxAccount", "value": _to_display(sbt_q)},
            {"label": "All other Tax Accounts - IWT,MFWT,FCWT & Others", "value": _to_display(other_accounts)},

            {"label": "BUSINESS ADDRESS ", "value": ""},
            {"label": "PhysicalAddressProvince", "value": _to_display(trm_map.get("physicaladdressprovince"))},
            {"label": "Residential Address", "value": _to_display(residential_address)},
            {"label": "Mailing Address", "value": _to_display(mailing_address)},

            {"label": "BUSINESS ACTIVITY", "value": ""},
            {"label": "SectorActivity", "value": _to_display(trm_map.get("sectoractivity"))},
            {"label": "EnterpriseActivity", "value": _to_display(trm_map.get("enterpriseactivity"))},
            {"label": "Enterprise Activity Number", "value": _to_display(trm_map.get("entactivitycode"))},

            {"label": "BUSINESS CONTACT DETAILS", "value": ""},
            {"label": "Contact Name", "value": _to_display(trm_map.get("entcontactname") or trm_map.get("contactname"))},
            {"label": "Contact title", "value": _to_display(trm_map.get("title"))},
            {"label": "Phone Number", "value": _to_display(phone_number)},
            {"label": "Email Address", "value": _to_display(trm_map.get("entcontemail") or trm_map.get("contactemail"))},
            {"label": "Representative Name", "value": _to_display(trm_map.get("repcontactname"))},
            {"label": "Representative Type", "value": _to_display(trm_map.get("reptype"))},
            {"label": "RepresentativePhone Number", "value": _to_display(rep_phone)},
            {"label": "Representative Email Address", "value": _to_display(trm_map.get("repcontemail"))},

            {"label": "TAX COMPLIANCE INDICATOR FOR MAIN TAX TYPES (lodgments/Tax Account balances)", "value": ""},
            {"label": "CIT Outstanding Returns", "value": _to_display(cit_outstanding)},
            {"label": "CIT Account Balance", "value": _to_display(cit_account_balance if cit_total_records else None)},
            {"label": "GST Outstanding Returns", "value": _to_display(gst_outstanding)},
            {"label": "GST Account Balance", "value": _to_display(gst_account_balance if gst_total_records else None)},
            {"label": "SWT Outstanding Returns", "value": _to_display(swt_outstanding)},
            {"label": "SWT Account Balance", "value": _to_display(swt_account_balance if swt_total_records else None)},

            {"label": "ASSETS & LIABILITIES", "value": ""},
            {"label": "Total Number of Employees (to be extracted on SWT data latest returns submitted)", "value": _to_display(employees_val)},
            {"label": "Annual Turnover (to be extracted from GST return)", "value": _to_display(annual_turnover if gst_total_records else None)},
            {"label": "Total Assets (to be extracted from CIT return)", "value": _to_display(total_assets)},
            {"label": "Total liabilities (to be extracted from CIT return)", "value": _to_display(total_liabilities)},

            {"label": "RISK ANALYSIS RESULT", "value": ""},
            {"label": "CIT Risk Identified", "value": _to_display(cit_risk)},
            {"label": "GST Risk Identified", "value": _to_display(gst_risk)},
            {"label": "SWT Risk Identified", "value": _to_display(swt_risk)},
            {"label": "Other Risk Identified", "value": _to_display(overall_risk)},

            {"label": "RECOMMENDATION", "value": ""},
            {"label": "CIT Recommendation", "value": _to_display(cit_reco)},
            {"label": "GST Recommendation", "value": _to_display(gst_reco)},
            {"label": "SWT Recommendation", "value": _to_display(swt_reco)},
            {"label": "Other Recommendation", "value": _to_display(other_reco)},
        ]

        if request.args.get("debug") == "1":
            print("structured_report:", structured_report[:5] if structured_report else "EMPTY")
            return jsonify({
                "structured_report": structured_report
            })

        response = {
            "gst": gst_response,
            "swt": swt_response,
            "structured_report": structured_report
        }

        # Optional debug validation (only when ?debug=1)
        if request.args.get("debug") == "1":
            def fmt_pairs(pairs):
                return [f"{y:04d}-{m:02d}" for (y, m) in sorted(pairs)]

            gst_actual = {(int(r.y), int(r.m)) for r in gst_actual_rows if getattr(r, "y", None) is not None and getattr(r, "m", None) is not None}
            swt_actual = {(int(r.y), int(r.m)) for r in swt_actual_rows if getattr(r, "y", None) is not None and getattr(r, "m", None) is not None}

            expected_list = fmt_pairs(expected_set)
            gst_actual_list = fmt_pairs(gst_actual)
            swt_actual_list = fmt_pairs(swt_actual)
            gst_missing_list = fmt_pairs(expected_set - gst_actual) if gst_total_records else []
            swt_missing_list = fmt_pairs(expected_set - swt_actual) if swt_total_records else []

            # CIT join coverage for this TIN
            cit_years_set = {int(r.y) for r in cit_year_rows if getattr(r, "y", None) is not None}
            ml_year_rows = db.session.execute(text("""
                SELECT DISTINCT tax_period_year AS y
                FROM cit_fraud_justification
                WHERE TRIM(tin) = TRIM(:tin)
                  AND tax_period_year BETWEEN :sy AND :ey
            """), {"tin": tin, "sy": start_y, "ey": end_y}).fetchall()
            ml_years_set = {int(r.y) for r in ml_year_rows if getattr(r, "y", None) is not None}
            matched_years = cit_years_set & ml_years_set
            join_coverage = (len(matched_years) / len(cit_years_set) * 100) if cit_years_set else 0

            swt_fallback_used = 0
            if swt_total_records:
                swt_latest_any = db.session.execute(text("""
                    SELECT employees_on_payroll, employees_paid_swt
                    FROM swt_fraud_justification
                    WHERE TRIM(CAST(tin AS CHAR(20))) = TRIM(:tin_norm)
                    ORDER BY tax_period_year DESC, tax_period_month DESC
                    LIMIT 1
                """), {"tin_norm": tin_norm}).fetchone()
                if swt_latest_any and swt_latest:
                    if (swt_latest_any[0] is None and swt_latest_any[1] is None) and (swt_latest[0] is not None or swt_latest[1] is not None):
                        swt_fallback_used = 1

            response["debug"] = {
                "gst_balance_primary": primary_balance,
                "gst_balance_fallback": fallback_balance,
                "gst_balance_mismatch_pct": round(gst_mismatch_pct, 2),
                "expected_months": expected_list,
                "gst_actual_months": gst_actual_list,
                "gst_missing_months": gst_missing_list,
                "swt_actual_months": swt_actual_list,
                "swt_missing_months": swt_missing_list,
                "cit_consistency_year": cit_max_year,
                "cit_join_coverage_pct": round(join_coverage, 2),
                "swt_employee_fallback_used": swt_fallback_used
            }

        return jsonify(response)

    except Exception as e:
        db.session.rollback()
        import traceback
        print("ERROR:", str(e))
        print(traceback.format_exc())
        return jsonify({"status": "error", "message": str(e)}), 500


# --------------------------------------------------------
#   EXPORT: CSV (structured_report)
# --------------------------------------------------------
@bp.route("/export-csv", methods=["OPTIONS"])
def export_csv_options():
    return ("", 200)


@bp.get("/export-csv")
@jwt_required()
def export_csv():
    data = _get_structured_report_from_summary()
    if isinstance(data, dict) and data.get("_error"):
        payload = data.get("payload") or {"error": "Export failed"}
        return jsonify(payload), data.get("status_code") or 400
    structured_report = data.get("structured_report") if isinstance(data, dict) else None
    if not structured_report:
        structured_report = []
    if not isinstance(structured_report, list):
        return jsonify({"status": "error", "message": "structured_report is not list"}), 500

    if _is_no_data(structured_report):
        si = io.StringIO()
        cw = csv.writer(si)
        cw.writerow(["No data exist for this taxpayer or TIN", ""])
        mem = io.BytesIO()
        mem.write(si.getvalue().encode("utf-8"))
        mem.seek(0)
        return send_file(mem, mimetype="text/csv", as_attachment=True, download_name="Risk_Profile_No_Data.csv")

    si = io.StringIO()
    cw = csv.writer(si)
    cw.writerow(["#NAME?", "VALUE"])
    for row in structured_report:
        cw.writerow([row.get("label", ""), row.get("value", "NA")])

    mem = io.BytesIO()
    mem.write(si.getvalue().encode("utf-8"))
    mem.seek(0)
    tin = request.args.get("tin") or "UNKNOWN"
    return send_file(mem, mimetype="text/csv", as_attachment=True, download_name=f"Risk_Profile_{tin}.csv")


# --------------------------------------------------------
#   EXPORT: EXCEL (structured_report)
# --------------------------------------------------------
@bp.route("/export-excel", methods=["OPTIONS"])
def export_excel_options():
    return ("", 200)


@bp.get("/export-excel")
@jwt_required()
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        return jsonify({"status": "error", "message": "openpyxl is required to export Excel"}), 500

    data = _get_structured_report_from_summary()
    if isinstance(data, dict) and data.get("_error"):
        payload = data.get("payload") or {"error": "Export failed"}
        return jsonify(payload), data.get("status_code") or 400
    structured_report = data.get("structured_report") if isinstance(data, dict) else None
    if not structured_report:
        structured_report = []
    if not isinstance(structured_report, list):
        return jsonify({"status": "error", "message": "structured_report is not list"}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Profiling"
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 40

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    sub_fill = PatternFill("solid", fgColor="E5E7EB")
    risk_high = PatternFill("solid", fgColor="DC2626")
    risk_low = PatternFill("solid", fgColor="16A34A")
    risk_na = PatternFill("solid", fgColor="9CA3AF")

    if _is_no_data(structured_report):
        ws.merge_cells("A1:B1")
        cell = ws["A1"]
        cell.value = "No data exist for this taxpayer or TIN"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        ws.append(["#NAME?", "VALUE"])
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws["B1"].font = header_font
        ws["B1"].fill = header_fill

        for row in structured_report:
            label = row.get("label", "")
            value = row.get("value", "NA")
            ws.append([label, value])
            current_row = ws.max_row

            if value in ("", None):
                ws[f"A{current_row}"].fill = header_fill
                ws[f"B{current_row}"].fill = header_fill
                ws[f"A{current_row}"].font = header_font
                ws[f"B{current_row}"].font = header_font
                continue

            if label and "Risk Identified" in label:
                risk = _risk_fill(value)
                if risk == "HIGH":
                    ws[f"B{current_row}"].fill = risk_high
                elif risk in ("MEDIUM", "LOW"):
                    ws[f"B{current_row}"].fill = risk_low
                else:
                    ws[f"B{current_row}"].fill = risk_na
                ws[f"B{current_row}"].font = header_font

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    tin = request.args.get("tin") or "UNKNOWN"
    return send_file(mem, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Risk_Profile_{tin}.xlsx")


# --------------------------------------------------------
#   EXPORT: PDF (structured_report)
# --------------------------------------------------------
@bp.route("/export-pdf", methods=["OPTIONS"])
def export_pdf_options():
    return ("", 200)


@bp.get("/export-pdf")
@jwt_required()
def export_pdf():
    tin = (request.args.get("tin") or "").strip()
    start_date = (request.args.get("start_date") or "").strip()
    end_date = (request.args.get("end_date") or "").strip()

    if not tin:
        return jsonify({"error": "TIN required"}), 400

    if not start_date or not end_date:
        return jsonify({"error": "Date range required"}), 400

    reportlab_available = True
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except Exception:
        reportlab_available = False

    data = _get_structured_report_from_summary()
    if isinstance(data, dict) and data.get("_error"):
        payload = data.get("payload") or {"error": "Export failed"}
        return jsonify(payload), data.get("status_code") or 400
    structured_report = data.get("structured_report") if isinstance(data, dict) else None
    if not structured_report:
        structured_report = []
    if not isinstance(structured_report, list):
        return jsonify({"status": "error", "message": "structured_report is not list"}), 500
    try:
        print("structured_report:", structured_report[:5] if structured_report else "EMPTY")
    except Exception:
        pass

    if not reportlab_available:
        lines = [
            f"Tax Compliance Report for TIN: {tin}",
            f"Period: {start_date} - {end_date}",
            f"Generated On: {datetime.now().strftime('%Y-%m-%d')}",
            "",
        ]
        if _is_no_data(structured_report):
            lines.append("No data exist for this taxpayer or TIN")
        else:
            for row in structured_report:
                label = row.get("label", "") or ""
                value = row.get("value", "NA")
                if value in ("", None):
                    lines.append("")
                    lines.append(label.upper())
                else:
                    lines.append(f"{label}: {value}")

        mem = io.BytesIO(_build_simple_pdf_bytes(lines))
        mem.seek(0)
        return send_file(
            mem,
            as_attachment=True,
            download_name=f"risk_profile_{tin}.pdf",
            mimetype="application/pdf"
        )

    mem = io.BytesIO()
    doc = SimpleDocTemplate(mem, pagesize=A4, topMargin=80, bottomMargin=40)
    styles = getSampleStyleSheet()
    elements = []

    generated_on = datetime.now().strftime("%Y-%m-%d")

    def _draw_header_footer(canvas, doc_obj):
        canvas.saveState()
        width, height = A4
        # Header
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(width / 2, height - 30, f"Tax Compliance Report for TIN: {tin}")
        canvas.setFont("Helvetica", 9)
        if start_date and end_date:
            canvas.drawCentredString(width / 2, height - 44, f"Period: {start_date} - {end_date}")
        canvas.drawCentredString(width / 2, height - 56, f"Generated On: {generated_on}")

        # Footer
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.HexColor("#6b7280"))
        canvas.drawRightString(width - 30, 20, f"Page {doc_obj.page}")
        canvas.restoreState()

    elements.append(Spacer(1, 6))

    if _is_no_data(structured_report):
        elements.append(Spacer(1, 50))
        elements.append(Paragraph("<b>No data exist for this taxpayer or TIN</b>", styles["Title"]))
        doc.build(elements, onFirstPage=_draw_header_footer, onLaterPages=_draw_header_footer)
        mem.seek(0)
        return send_file(mem, mimetype="application/pdf", as_attachment=True, download_name=f"risk_profile_{tin}.pdf")

    for idx, row in enumerate(structured_report):
        label = row.get("label", "") or ""
        value = row.get("value", "NA")

        # Section header (grey with left blue border)
        if value in ("", None):
            header_tbl = Table([[label.upper()]], colWidths=[500], hAlign="LEFT")
            header_style = TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#e5e7eb")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("INNERGRID", (0, 0), (-1, -1), 0, colors.white),
                ("BOX", (0, 0), (-1, -1), 0, colors.white),
                ("LINEBEFORE", (0, 0), (0, 0), 4, colors.HexColor("#6366f1")),
            ])
            header_tbl.setStyle(header_style)
            elements.append(header_tbl)
            elements.append(Spacer(1, 6))
            continue

        # Normal row (label/value)
        is_risk = "Risk Identified" in label
        display_value = Paragraph(str(value), styles["Normal"])
        row_tbl = Table([[label, display_value]], colWidths=[320, 180], hAlign="LEFT")
        row_style = TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("INNERGRID", (0, 0), (-1, -1), 0, colors.white),
            ("BOX", (0, 0), (-1, -1), 0, colors.white),
            ("LINEBELOW", (0, 0), (-1, -1), 0.35, colors.HexColor("#e5e7eb")),
        ])

        if idx % 2 == 0:
            row_style.add("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f9fafb"))

        if is_risk:
            risk = _risk_fill(value)
            if risk == "HIGH":
                row_style.add("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#dc2626"))
                row_style.add("TEXTCOLOR", (1, 0), (1, 0), colors.white)
                row_style.add("FONTNAME", (1, 0), (1, 0), "Helvetica-Bold")
            elif risk in ("MEDIUM", "LOW"):
                row_style.add("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#16a34a"))
                row_style.add("TEXTCOLOR", (1, 0), (1, 0), colors.white)
                row_style.add("FONTNAME", (1, 0), (1, 0), "Helvetica-Bold")
            else:
                row_style.add("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#9ca3af"))
                row_style.add("TEXTCOLOR", (1, 0), (1, 0), colors.white)

        row_tbl.setStyle(row_style)
        elements.append(row_tbl)
        elements.append(Spacer(1, 6))

    doc.build(elements, onFirstPage=_draw_header_footer, onLaterPages=_draw_header_footer)
    mem.seek(0)
    return send_file(mem, mimetype="application/pdf", as_attachment=True, download_name=f"risk_profile_{tin}.pdf")

